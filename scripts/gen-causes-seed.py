#!/usr/bin/env python3
"""
从用户桌面 xlsx 生成 prisma/seeds/causes-{civil,criminal,administrative}.ts

xlsx 路径：~/Desktop/民事刑事行政诉讼案由分级分类表.xlsx
"""
import os
import re
from pathlib import Path
from openpyxl import load_workbook

XLSX = Path.home() / "Desktop" / "民事刑事行政诉讼案由分级分类表.xlsx"
OUT_DIR = Path(__file__).parent.parent / "prisma" / "seeds"

# 中文小写数字到阿拉伯
CN_DIGIT = {"零": 0, "一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9}

def cn_to_int(s):
    """支持 一~九十九：'十' / '十一' / '二十' / '二十一' 等"""
    if not s:
        return None
    s = s.strip()
    if s in CN_DIGIT:
        return CN_DIGIT[s]
    if "十" in s:
        parts = s.split("十")
        if len(parts) == 2:
            tens = CN_DIGIT[parts[0]] if parts[0] else 1
            ones = CN_DIGIT[parts[1]] if parts[1] else 0
            return tens * 10 + ones
    return None

def parse_part(s):
    """'第一部分' → 1"""
    if not s:
        return None
    m = re.match(r"第([一二三四五六七八九十]+)部分", s)
    if not m:
        return None
    return cn_to_int(m.group(1))

def parse_cn(s):
    """'一' / '（一）' / '二十一' → 1/21"""
    if not s:
        return None
    s = s.strip().strip("（）()")
    return cn_to_int(s)

def parse_int(s):
    if s is None:
        return None
    if isinstance(s, str):
        # 兼容 "（1）" "(1)" "1." 等格式
        cleaned = re.sub(r"[（）()\s\.、]", "", s)
        if cleaned.isdigit():
            return int(cleaned)
    try:
        return int(s)
    except (ValueError, TypeError):
        return None

def esc(s):
    """TS string escape"""
    return (s or "").replace("\\", "\\\\").replace('"', '\\"')

def emit_ts(items, out_path, header, export_name):
    lines = [header, "", f"export const {export_name} = ["]
    for it in items:
        parts = [f'code: "{it["code"]}"', f'name: "{esc(it["name"])}"', f'level: {it["level"]}']
        if it.get("parentCode"):
            parts.append(f'parentCode: "{it["parentCode"]}"')
        if it.get("shortName"):
            parts.append(f'shortName: "{esc(it["shortName"])}"')
        if it.get("keywords"):
            kw = ", ".join(f'"{esc(k)}"' for k in it["keywords"])
            parts.append(f'keywords: [{kw}]')
        lines.append(f'  {{ {", ".join(parts)} }},')
    lines.append("];")
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"  → {out_path.name} ({len(items)} 条)")

def gen_civil(wb):
    """民事案由：1-4 级"""
    ws = wb["民事案由"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    items = []
    seen = set()
    for r in rows[1:]:
        d = dict(zip(headers, r))
        lvl = d.get("层级")
        l1n, l1m = parse_part(d.get("一级编号")), d.get("一级案由")
        l2n, l2m = parse_cn(d.get("二级编号")), d.get("二级案由")
        l3n, l3m = parse_int(d.get("三级编号")), d.get("三级案由")
        l4n, l4m = parse_int(d.get("四级编号")), d.get("四级案由")
        if l1n is None:
            continue
        code1 = f"CC-{l1n}"
        if code1 not in seen and l1m:
            items.append({"code": code1, "name": l1m.strip(), "level": 1})
            seen.add(code1)
        if lvl == 1:
            continue
        if l2n is None:
            continue
        code2 = f"CC-{l1n}-{l2n}"
        if code2 not in seen and l2m:
            items.append({"code": code2, "name": l2m.strip(), "level": 2, "parentCode": code1})
            seen.add(code2)
        if lvl == 2:
            continue
        if l3n is None:
            continue
        code3 = f"CC-{l1n}-{l2n}-{l3n}"
        if code3 not in seen and l3m:
            items.append({"code": code3, "name": l3m.strip(), "level": 3, "parentCode": code2})
            seen.add(code3)
        if lvl == 3:
            continue
        if l4n is None:
            continue
        code4 = f"CC-{l1n}-{l2n}-{l3n}-{l4n}"
        if code4 not in seen and l4m:
            items.append({"code": code4, "name": l4m.strip(), "level": 4, "parentCode": code3})
            seen.add(code4)
    header = '/**\n * 民事案件案由 — 完整版（基于《民事案件案由规定》2025 修正）\n * 由 scripts/gen-causes-seed.py 从用户桌面 xlsx 生成\n */\n\n/* eslint-disable */'
    emit_ts(items, OUT_DIR / "causes-civil.ts", header.replace("/* eslint-disable */", ""), "civilCauses")

def gen_criminal(wb):
    """刑事罪名：章/节/罪名"""
    ws = wb["刑事罪名"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    items = []
    seen = set()
    # 给章和节分配序号
    chap_idx = {}
    sect_idx = {}  # (chap, sect_name) -> int
    chap_count = 0
    for r in rows[1:]:
        d = dict(zip(headers, r))
        chap = d.get("刑法分则章")
        if chap and chap not in chap_idx:
            chap_count += 1
            chap_idx[chap] = chap_count
            code = f"CR-{chap_count}"
            items.append({"code": code, "name": chap.strip(), "level": 1})
            seen.add(code)
        sect = d.get("刑法分则节")
        if sect:
            key = (chap, sect)
            if key not in sect_idx:
                # 节序号 = 在该章下第几个 unique 节
                sect_count = sum(1 for k in sect_idx if k[0] == chap) + 1
                sect_idx[key] = sect_count
                code2 = f"CR-{chap_idx[chap]}-{sect_count}"
                items.append({
                    "code": code2,
                    "name": sect.strip(),
                    "level": 2,
                    "parentCode": f"CR-{chap_idx[chap]}"
                })
                seen.add(code2)
    # 第二遍写罪名
    crime_count_per_parent = {}
    for r in rows[1:]:
        d = dict(zip(headers, r))
        crime = d.get("罪名")
        if not crime:
            continue
        chap = d.get("刑法分则章")
        sect = d.get("刑法分则节")
        if sect:
            parent_code = f"CR-{chap_idx[chap]}-{sect_idx[(chap, sect)]}"
        elif chap:
            parent_code = f"CR-{chap_idx[chap]}"
        else:
            continue
        n = crime_count_per_parent.get(parent_code, 0) + 1
        crime_count_per_parent[parent_code] = n
        code3 = f"{parent_code}-{n}"
        items.append({
            "code": code3,
            "name": crime.strip(),
            "level": 3,
            "parentCode": parent_code,
        })
    header = '/**\n * 刑事罪名 — 完整版 484 项（基于刑法分则 + 最高法补充规定）\n * 由 scripts/gen-causes-seed.py 从用户桌面 xlsx 生成\n */'
    emit_ts(items, OUT_DIR / "causes-criminal.ts", header, "criminalCauses")

def gen_admin(wb):
    """行政案由：1-3 级"""
    ws = wb["行政案由"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    items = []
    seen = set()
    # 行政表没有阿拉伯一级编号，用"行政行为"作为唯一一级
    # 实际看数据，所有 行 的一级案由都是同一个：先用 "行政行为" / "行政复议" 等做唯一性
    l1_idx = {}
    for r in rows[1:]:
        d = dict(zip(headers, r))
        l1m = d.get("一级案由")
        if l1m and l1m not in l1_idx:
            l1_idx[l1m] = len(l1_idx) + 1
            code = f"AD-{l1_idx[l1m]}"
            if code not in seen:
                items.append({"code": code, "name": l1m.strip(), "level": 1})
                seen.add(code)
    # 第二遍：二级
    l2_idx = {}
    for r in rows[1:]:
        d = dict(zip(headers, r))
        l1m = d.get("一级案由")
        l2m = d.get("二级案由")
        if not (l1m and l2m):
            continue
        key = (l1m, l2m)
        if key not in l2_idx:
            sub = sum(1 for k in l2_idx if k[0] == l1m) + 1
            l2_idx[key] = sub
            code = f"AD-{l1_idx[l1m]}-{sub}"
            if code not in seen:
                items.append({
                    "code": code,
                    "name": l2m.strip(),
                    "level": 2,
                    "parentCode": f"AD-{l1_idx[l1m]}",
                })
                seen.add(code)
    # 第三遍：三级
    l3_count = {}
    for r in rows[1:]:
        d = dict(zip(headers, r))
        l1m = d.get("一级案由")
        l2m = d.get("二级案由")
        l3m = d.get("三级案由")
        if not (l1m and l2m and l3m):
            continue
        parent = f"AD-{l1_idx[l1m]}-{l2_idx[(l1m, l2m)]}"
        n = l3_count.get(parent, 0) + 1
        l3_count[parent] = n
        # 三级名形如 "1.警告" / "2.通报批评"，去掉前缀编号
        clean = re.sub(r"^\d+[\.、]\s*", "", l3m.strip())
        code = f"{parent}-{n}"
        items.append({
            "code": code,
            "name": clean,
            "level": 3,
            "parentCode": parent,
        })
    header = '/**\n * 行政案件案由 — 完整版（基于最高法《关于行政案件案由的暂行规定》2021）\n * 由 scripts/gen-causes-seed.py 从用户桌面 xlsx 生成\n */'
    emit_ts(items, OUT_DIR / "causes-administrative.ts", header, "administrativeCauses")

def main():
    print(f"读取 {XLSX}")
    wb = load_workbook(XLSX, read_only=True)
    gen_civil(wb)
    gen_criminal(wb)
    gen_admin(wb)
    print("完成")

if __name__ == "__main__":
    main()
